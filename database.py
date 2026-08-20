import subprocess
import sys
import re
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent

MAIN_FILE = BASE_DIR / "wellness_monitor.py"
SESSION_FILE = BASE_DIR / "wellness_session.txt"
DATABASE_FILE = BASE_DIR / "wellness_database.xlsx"

HEADERS = [
    "User / Candidate Name",
    "Reference Number",
    "Date",
    "Start Time",
    "End Time",
    "Session Duration (sec)",
    "Face Width (px)",
    "Distance",
    "Sitting Time (sec)",
    "Break Countdown (sec)",
    "Wellness Score",
    "Recommendation",
    "Health Tip",
    "Session Status"
]


def create_database():

    if DATABASE_FILE.exists():
        return

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Wellness Sessions"
    sheet.append(HEADERS)

    fill = PatternFill(
        "solid",
        fgColor="1F4E78"
    )

    font = Font(
        bold=True,
        color="FFFFFF"
    )

    for cell in sheet[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    widths = [
        28, 22, 18, 14, 14, 24,
        18, 18, 22, 24, 20, 40, 45, 20
    ]

    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[
            get_column_letter(i)
        ].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    workbook.save(DATABASE_FILE)


def ask_details():

    print("\n" + "=" * 60)
    print("          AMBIENT AI WELLNESS MONITOR")
    print("                USER LOGIN")
    print("=" * 60)

    while True:
        name = input(
            "\nEnter User / Candidate Name: "
        ).strip()

        if name:
            break

        print("Name cannot be empty.")

    while True:
        reference = input(
            "Enter Reference Number: "
        ).strip()

        if reference:
            break

        print("Reference number cannot be empty.")

    return name, reference


def read_session():

    if not SESSION_FILE.exists():
        return {}

    text = SESSION_FILE.read_text(
        encoding="utf-8"
    )

    data = {}

    for line in text.splitlines():

        if ":" not in line:
            continue

        key, value = line.split(
            ":",
            1
        )

        data[key.strip()] = value.strip()

    return data


def value(data, key):

    return data.get(key, "")


def save_session(
    name,
    reference,
    start_time,
    end_time,
    data,
    status
):

    workbook = openpyxl.load_workbook(
        DATABASE_FILE
    )

    sheet = workbook["Wellness Sessions"]

    duration = int(
        (end_time - start_time).total_seconds()
    )

    sheet.append([
        name,
        reference,
        start_time.strftime("%Y-%m-%d"),
        start_time.strftime("%H:%M:%S"),
        end_time.strftime("%H:%M:%S"),
        duration,

        value(data, "Face Width"),
        value(data, "Distance"),

        value(data, "Sitting Time"),
        value(data, "Break Countdown"),

        value(data, "Wellness Score"),

        value(data, "Recommendation"),
        value(data, "Health Tip"),

        status
    ])

    for cell in sheet[sheet.max_row]:
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=True
        )

    sheet.auto_filter.ref = sheet.dimensions

    workbook.save(DATABASE_FILE)


def main():

    if not MAIN_FILE.exists():

        print(
            "\nERROR: wellness_monitor.py "
            "was not found."
        )

        input(
            "\nPress Enter to exit..."
        )

        return

    create_database()

    name, reference = ask_details()

    # Remove old session data so a previous candidate
    # can never be accidentally saved.
    if SESSION_FILE.exists():

        try:
            SESSION_FILE.unlink()

        except PermissionError:

            print(
                "\nClose wellness_session.txt "
                "and try again."
            )

            input(
                "\nPress Enter to exit..."
            )

            return

    start_time = datetime.now()

    print("\n" + "=" * 60)
    print(f"User      : {name}")
    print(f"Reference : {reference}")
    print(
        f"Started   : "
        f"{start_time.strftime('%H:%M:%S')}"
    )
    print("=" * 60)

    print(
        "\nStarting Wellness Monitor..."
    )

    print(
        "Press Q in the Wellness Monitor "
        "window when finished."
    )

    try:

        result = subprocess.run(
            [
                sys.executable,
                str(MAIN_FILE)
            ],
            cwd=str(BASE_DIR)
        )

    except Exception as error:

        print(
            f"\nCould not start Wellness Monitor:"
            f"\n{error}"
        )

        input(
            "\nPress Enter to exit..."
        )

        return

    end_time = datetime.now()

    data = read_session()

    if not data:

        print(
            "\nWARNING: No wellness session data "
            "was created."
        )

        print(
            "Make sure you finish the monitor "
            "by pressing Q."
        )

        input(
            "\nPress Enter to exit..."
        )

        return

    status = (
        "Completed"
        if result.returncode == 0
        else f"Ended with code {result.returncode}"
    )

    save_session(
        name,
        reference,
        start_time,
        end_time,
        data,
        status
    )

    print("\n" + "=" * 60)
    print("       WELLNESS SESSION SAVED")
    print("=" * 60)

    print(f"User           : {name}")
    print(f"Reference      : {reference}")
    print(
        f"Wellness Score : "
        f"{value(data, 'Wellness Score')}"
    )
    print(
        f"Excel          : "
        f"{DATABASE_FILE}"
    )

    input(
        "\nPress Enter to close..."
    )


if __name__ == "__main__":
    main()